#Contains: get_plybook, append_mat_properties, count_laminate_materials
#http://www.youlikeprogramming.com/2012/03/examples-reading-excel-xlsx-2007-documents-using-python-openpyxl/
import MECH530_matrices #Computes the Q_on/off, S_on/off matrices

import openpyxl #Allows me to read from a .xlsx file


def getplybook():
	#Opens .xlsx raw plybook, reads data in the range specified (max 24 plies + core) and outputs arranged plybook data
	workbook = openpyxl.load_workbook(filename = 'MECH530_Material_Database_ElaineCraigie.xlsx', use_iterators = True, data_only = True)

	plybook_sheet = workbook.worksheets[0] #True provided that Plybook remains the first sheet in the file
	plybook_range = plybook_sheet.iter_rows('C4:F43') #contains a list of material plies in the laminate

	#PLYBOOK--------------------------------------------------------------------------------------------------------------------------------------
	raw_plybook = [] #This list will contain all information provided in the Plybook and sort out doubled plies, etc.
	raw_plybook = [[cell.value for cell in row] for row in plybook_range]

	#Going to need to count the number of rows in my list (this will later have to be corrected is a ply is doubled)
	plybook = [[[] for i in range(19)] for j in range(len(raw_plybook))] #Initiatilizing: have 24 rows and 13 columns filled with blanks
	raw_r = 0 #row counter (for raw plybook) 
	ply_r = 0 #row counter (for final plybook) #Will count our total number of plies in the system (including the CORE) and remove the additional empty rows
	h_laminate = 0 #will tally the total thickness of my laminate
	h_core = 0 #checks for a core (if no core, returns zero thickness)
	while raw_r <= 39:
		if raw_plybook[raw_r][0] != 0:
			# print "-----------------"
			# print "ply_r = %d" %ply_r
			# print "raw_plybook val = %d" % raw_plybook[raw_r][0]
			if raw_plybook[raw_r][1] == 'CORE': #Check if there is a CORE
				h_core = raw_plybook[raw_r][3] #assigns a core thickness (if there is one)
			#Begin transfering values from raw_plybook to plybook	
			plybook[ply_r][0] = ply_r	
			c = 1 #my column counter
			while c <= 3:
				plybook[ply_r][c] = raw_plybook[raw_r][c] #assign cell value from raw_plybook to plybook
				c += 1 #increment column counter
			h_laminate += plybook[ply_r][3]	#add thickness of new ply
			if raw_plybook[raw_r][0] > 1:
				repeats = raw_plybook[raw_r][0] - 1
				# print "repeats = %d" % repeats
				repeats_counter = 1
				while repeats_counter <= repeats:
					ply_r += 1 #increment Plybook ply counter
					# print "ply_r = %d" %ply_r
					plybook[ply_r][0] = ply_r
					c = 1 #reset column counter
					while c <= 3:
						plybook[ply_r][c] = raw_plybook[raw_r][c]
						c += 1 #increment column counter
					repeats_counter += 1 #increment repeats counter
					h_laminate += plybook[ply_r][3] #add thickness of new ply
			raw_r += 1
			ply_r += 1		
		else:
			break

	plies = ply_r #this is the total number of plies in the laminate
	# print "plies = %d" %plies
	# print "h_laminate = %f" %round(h_laminate,3)
	# print "h_core = %f" %round(h_core,3)
	return plies, h_laminate, h_core, plybook

def append_material_properties(plies, plybook):
	#Opens .xlsx material properties data and appends material property list based on materials contained in plybook. Returns plybook
	workbook = openpyxl.load_workbook(filename = 'MECH530_Material_Database_ElaineCraigie.xlsx', use_iterators = True, data_only = True)

	property_sheet = workbook.worksheets[1] #True provided that Properties remains the second sheet in the file
	property_range = property_sheet.iter_rows('A10:M17') #contains all of my materials and their properties

	#MATERIAL DATABASE---------------------------------------------------------------------------------------------------------------------------
	mat_properties = [] #This list will contain all information provided in the Plybook and sort out doubled plies, etc.
	mat_properties = [[cell.value for cell in row] for row in property_range]

	#Build these properties into my plybook list so that each row has all of the appropriate material properties associated with it
	ply_r = 0

	while ply_r <= plies - 1:
		mat_r = 0 #counts my rows in the mat_properties list (5 materials so far - change limit on WHILE loop if this evolves)
		while mat_r <= 6:
			# print 'ply_r_2 = ' + str(ply_r_2)
			# print 'mat_r = ' + str(mat_r)
			if plybook[ply_r][1] == mat_properties[mat_r][0]:
				mat_c = 4 #counts my columns in the mat_properties list (these are the correct indices for the plybook list, but are +3 for the mat_properties list)
				while mat_c <= 12: #Add material properties to each ply in Plybook
					plybook[ply_r][mat_c] = mat_properties[mat_r][mat_c - 1]
					mat_c += 1
				#The following line builds the S_on, Q_on, S_off, Q_off, Uq, Us matrices and vector into plybook for EACH ply respectively	
				plybook = MECH530_matrices.get_on_off_matrices(plies, plybook, ply_r)	
			mat_r += 1		
		ply_r += 1

	return plybook


def count_laminate_materials(plies, plybook):
	#Check for all of the materials in my laminate to print properties PER material
	materials_used = {} #define empty dictionary which will eventually contain the material(s) used
	ply_r = 0 
	num_materials = 1
	materials_used[num_materials] = plybook[ply_r][1] #initialize the dictionary using the first material used
	ply_r = 1

	while ply_r < plies:
		check = 0
		mat_c = 1
		while mat_c <= num_materials:
			num_materials_counter = 1
			while num_materials_counter <= num_materials:
				if plybook[ply_r][1] == materials_used[num_materials_counter]:
					check += 1
				num_materials_counter += 1	
			mat_c += 1		
		if check == 0 and plybook[ply_r][1] != 'CORE': #Don't count core as another material
			num_materials += 1
			materials_used[num_materials] = plybook[ply_r][1]
		ply_r += 1

	return materials_used, num_materials	